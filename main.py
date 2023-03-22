from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl
import warnings
from PIL import Image
import re
import os
import glob
from datetime import datetime

from email_writer import EmailWriter

warnings.filterwarnings('ignore')

email_a1 = 'teo.ljubicic@a1.hr'
googleSheet = r'https://a1g.sharepoint.com/:x:/r/sites/o365ESD-ESMkoordinacija/Shared%20Documents/General/ITS%20GA,%20%C4%8Deka%20ESD.xlsx?d=wccefade660e746668e60a09a2ca476d6&csf=1&web=1&e=hcR8NL'

# OPEN CHROME
power_bi = 'https://app.powerbi.com/groups/a5cf286e-7692-4fc7-ad6c-db7e8f6c1fb5/reports/863cd96b-3344-4439-b442-3a2968f84e5a'
path = r'C:\Users\tljubicic\OneDrive - A1 Group\Dokumente\chromedriver.exe'

driver = webdriver.Chrome(path)
driver.get(power_bi)
driver.maximize_window()

# LOG IN - OK
driver.implicitly_wait(5)
driver.find_element_by_id("email").send_keys(email_a1)
driver.find_element_by_id("submitBtn").click()

driver.implicitly_wait(3)
driver.find_element_by_class_name("table-row").click()
print('Log in successful')

# VERIFICATION CODE - OK
sms_input = input('Enter verification code: ')
driver.find_element_by_id("idTxtBx_SAOTCC_OTC").send_keys(sms_input)
driver.find_element_by_id("idSubmit_SAOTCC_Continue").click()
driver.implicitly_wait(10)
time.sleep(3)

while True:
    if driver.find_elements_by_xpath('//mat-action-list[@class]//button[1]'):
        break
    else:
        print('Verification failed, try again')
        driver.find_element_by_id("idTxtBx_SAOTCC_OTC").clear()
        sms_input = input('Enter verification code: ')
        driver.find_element_by_id("idTxtBx_SAOTCC_OTC").send_keys(sms_input)
        driver.find_element_by_id("idSubmit_SAOTCC_Continue").click()
        time.sleep(1.5)
        continue

print('Verification successful')

# PIPELINE EXTRACTION - OK
driver.implicitly_wait(10)
driver.find_element_by_xpath('//mat-action-list[@class]//button[1]').click()
time.sleep(3)
driver.implicitly_wait(10)
statusi = driver.find_elements_by_xpath('//div[contains(@class,"pivotTableContainer")]//div[contains(@class,"pivotTable")]//div[contains(@class,"bodyCells")]//div[contains(@class,"pivot")]')

statusi_count = []
for num in statusi:
    a = num.text
    statusi_count.append(a)

spojeno_ceka_esm = statusi_count[1]
spojeno_ceka_esd = statusi_count[2]
potencijalni_linkovi = int(spojeno_ceka_esm) + int(spojeno_ceka_esd)
print('Pipeline extraction successful')


# USLUGE SCREENSHOT - OK
driver.find_element_by_xpath('//div[@title="SPOJENO, ČEKA ESD"]').click()

action = ActionChains(driver)
action.move_to_element(driver.find_element_by_xpath('//div[@title="SPOJENO, ČEKA ESD"]')).perform()
action.context_click().perform()
action.click(driver.find_element_by_xpath('//pbi-menu[@role="menu"]//button[contains(@title,"Drill through")]')).perform()

driver.implicitly_wait(10)
action.click(driver.find_element_by_xpath('//pbi-menu//button[@title="Detaljniji Status"]')).perform()

time.sleep(2)
driver.save_screenshot('usluge_screen.png')
print('Screenshot successful')

# CROPPING THE IMAGE - OK
im = Image.open('usluge_screen.png')

left = 1340
top = 265
right = 1875
bottom = 520

image_cropped = im.crop((left,top,right,bottom))

image_cropped = image_cropped.save('usluge_cropped.png')
print('Cropping successful')

# PRACENJE REALIZACIJE EXTRACITON -OK
driver.find_element_by_xpath('//mat-action-list[@class]//button[3]').click()
time.sleep(2)
total_ga_list = driver.find_elements_by_css_selector('#pvExplorationHost > div > div > exploration > div > explore-canvas > div > div.canvasFlexBox > div > div.displayArea.disableAnimations.fitToPage > div.visualContainerHost.visualContainerOutOfFocus > visual-container-repeat > visual-container:nth-child(6) > transform > div > div.visualContent > div > visual-modern > div > svg > svg > g.axisGraphicsContext.columnChart.lineChart > g.columnChartUnclippedGraphicsContext > svg > g:nth-child(1) > rect')
driver.implicitly_wait(10)
webdriver.ActionChains(driver).move_to_element(total_ga_list[-1]).click(total_ga_list[-1]).perform()

time.sleep(2)
total_ga = driver.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container[5]/transform/div/div[2]/div/visual-modern/div/div/div[2]/div[1]/div[6]/div/div/div/div').text

time.sleep(1)
webdriver.ActionChains(driver).move_to_element(total_ga_list[-2]).click(total_ga_list[-2]).perform()
driver.implicitly_wait(10)

prethodni_mjesec = driver.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container[5]/transform/div/div[2]/div/visual-modern/div/div/div[2]/div[1]/div[6]/div/div/div/div').text

print('Pracenje realizacije extraction successful')

# ESTIMACIJA REALIZACIJE EXTRACTION - OK
driver.find_element_by_xpath('//mat-action-list[@class]//button[4]').click()
time.sleep(2)
predikcija_ga = driver.find_element_by_css_selector('#pvExplorationHost > div > div > exploration > div > explore-canvas > div > div.canvasFlexBox > div > div.displayArea.disableAnimations.fitToPage > div.visualContainerHost.visualContainerOutOfFocus > visual-container-repeat > visual-container:nth-child(3) > transform > div > div.visualContent > div > visual-modern > div > svg > g.animatedNumber > text').text
print('Estimacija realizacije extraction successful')

## ITS GA EXTRACTION FROM EXTERNAL FILE - OK
driver.execute_script("window.open('');")
time.sleep(3)
driver.switch_to.window(driver.window_handles[1])
driver.get(googleSheet)
time.sleep(3)

driver.implicitly_wait(5)

# SWITCHING FRAMES
frame = driver.find_element_by_xpath('//iframe[@name="WebApplicationFrame"]')
driver.switch_to.frame(frame)
time.sleep(2)

# NAVIGATING THE MENU BAR
driver.find_element_by_xpath('//div[@id="FileMenuLauncherContianer"]').click()
time.sleep(2)
driver.implicitly_wait(10)
driver.find_element_by_xpath('//button[@id="FileSaveAsPage"]').click()
time.sleep(2)
driver.implicitly_wait(10)
driver.find_element_by_xpath('//button[@id="DownloadACopy"]').click()
time.sleep(2)

# EXTRACTING FROM EXCEL FILE
rootdir = r'C:\Users\tljubicic\Downloads'
regex = re.compile('ITS GA, čeka ESD.*')

files = glob.glob(rootdir+"\*.xlsx")

latest = max(files, key=os.path.getctime)

wb_obj = openpyxl.load_workbook(latest, data_only=True)
sheet_obj = wb_obj.active

manje_15 = 0
vise_15 = 0

for i in range(2, sheet_obj.max_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=26)
    if cell_obj.value != None:
        if int(cell_obj.value) <= 15:
            manje_15 = manje_15 + 1
        else:
            vise_15 = vise_15 + 1
    else:
        pass

print('ITS GA Excel extraction successful')

# EMAIL WRITER
writer = EmailWriter(total_ga,potencijalni_linkovi,spojeno_ceka_esd, manje_15, vise_15,predikcija_ga,spojeno_ceka_esm,prethodni_mjesec)

danasnjiDatum = datetime.today()
prviDan = datetime(danasnjiDatum.year,danasnjiDatum.month,1)
protekloDana = int(str(danasnjiDatum - prviDan).split(" days")[0])

if protekloDana <= 7:
    writer.PocetakMjeseca()
elif 7 < protekloDana <= 15:
    writer.PrvaPolovina()
else:
    writer.DrugaPolovina()

print('')
print('All tasks successful')
